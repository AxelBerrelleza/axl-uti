from .morning_star import *
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logging.basicConfig(filename='debug.log', level=logging.DEBUG)
# logging.basicConfig(level=logging.INFO, handlers={
#     logging.FileHandler('debug.log'),
#     logging.StreamHandler()
# })
logger = logging.getLogger(__name__)

class SpreadSheetComparation:
    """
    Populates a spreedsheet file using Morninstar data.
    Made for fundamental analysis
    """
    
    path: str = 'assets/base-sheet.xlsx'
    workbook: Workbook = None
    performanceIds: list
    symbols: list
    outputFilename: str = 'comparation.xlsx'
    initialColumn: int = 3
    initialRow: int = 2
    rowMap: dict = {
        'currentRatio': 3,
        'quickRatio': 4,
        'debtToEquity': 6,
        'inventoryTurnover': 7,
        'daysInventory': 8,
        'AssetTurnover': 9,
        'roe': 10,
        'netMargin': 11,
        'PER': 12,
        'PCF': 13,
        'PS': 14,
        'PBV': 15,
        'price': 18,
        'PER-5yr': 26,
        'PCF-5yr': 29,
        'PS-5yr': 32,
        'PBV-5yr': 35,
    }

    def __init__(self):
        print("Loading...")
        self.workbook = load_workbook(filename=self.path)

    def do(self):
        sheet: Worksheet = self.workbook.active
        try:
            self._loadSymbolsAsHeaders(sheet)
            self._loadOverviewData(sheet)
            self._loadInstrumentsPrice(sheet)
            self._loadPastAvgValuation(sheet)

        except Exception as ex:
            logger.error(f"{ ex }")

        self.workbook.save(filename=self.outputFilename)
        print("Finished")

    def _loadSymbolsAsHeaders(self, sheet: Worksheet):
        titleRow = 2
        for key, symbol in enumerate(self.symbols):
            sheet.cell(row=titleRow, column=self.initialColumn + key).value = symbol
    
    def _loadOverviewData(self, sheet: Worksheet):
        for key, perId in enumerate(self.performanceIds):
            response = getOverview(perId)

            sheet.cell(
                row=self.rowMap['currentRatio'], 
                column=self.initialColumn + key
            ).value = response['financialHealth']['currentRatio']
            sheet.cell(
                row=self.rowMap['quickRatio'], 
                column=self.initialColumn + key
            ).value = response['financialHealth']['quickRatio']
            sheet.cell(
                row=self.rowMap['debtToEquity'], 
                column=self.initialColumn + key
            ).value = response['financialHealth']['debtToEquity']

            sheet.cell(
                row=self.rowMap['roe'], 
                column=self.initialColumn + key
            ).value = response['efficiencyRatio']['returnOnEquity']
            sheet.cell(
                row=self.rowMap['netMargin'], 
                column=self.initialColumn + key
            ).value = response['profitabilityRatio']['netMargin']

            sheet.cell(
                row=self.rowMap['PER'], 
                column=self.initialColumn + key
            ).value = response['valuationRatio']['priceToEPS']
            sheet.cell(
                row=self.rowMap['PCF'], 
                column=self.initialColumn + key
            ).value = response['valuationRatio']['priceToCashFlow']
            sheet.cell(
                row=self.rowMap['PS'], 
                column=self.initialColumn + key
            ).value = response['valuationRatio']['priceToSales']
            sheet.cell(
                row=self.rowMap['PBV'], 
                column=self.initialColumn + key
            ).value = response['valuationRatio']['priceToBook']

    def _loadInstrumentsPrice(self, sheet: Worksheet):
        response = getInstrumentsPrice(self.symbols)        
        for key, data in enumerate(response):
            sheet.cell(
                row=self.rowMap['price'],
                column=self.initialColumn + key
            ).value = data['lastPrice']

    def _loadPastAvgValuation(self, sheet: Worksheet):
        PS_index = 0
        PER_index = 1
        PCF_index = 2
        PBV_index = 3
        def getFiveYearValue(index, rows):
            pastData: list = rows[index]['datum'] 
            penultimateIndex = len(pastData) - 2
            val = pastData[penultimateIndex]
            try:
                return float(val)
            except (ValueError, TypeError):
                return None
        
        for key, perId in enumerate(self.performanceIds):
            response = getAvgValuation(perId)
            resp_rows = response['Collapsed']['rows']

            sheet.cell(
                row=self.rowMap['PER-5yr'],
                column=self.initialColumn + key
            ).value = getFiveYearValue(PER_index, resp_rows)
            sheet.cell(
                row=self.rowMap['PCF-5yr'],
                column=self.initialColumn + key
            ).value = getFiveYearValue(PCF_index, resp_rows)
            sheet.cell(
                row=self.rowMap['PS-5yr'],
                column=self.initialColumn + key
            ).value = getFiveYearValue(PS_index, resp_rows)
            sheet.cell(
                row=self.rowMap['PBV-5yr'],
                column=self.initialColumn + key
            ).value = getFiveYearValue(PBV_index, resp_rows)