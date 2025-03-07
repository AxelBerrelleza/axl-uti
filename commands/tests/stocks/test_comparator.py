import pytest
from typer.testing import CliRunner
from logging import getLogger
from commands.stocks import stocks_app, Endpoints
from commands.tests.fixtures import APIResponses
from commands.stocks.SpreedSheetComparation import SpreadSheetComparation
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

runner = CliRunner()
logger = getLogger(__name__)

def test_help():
    result = runner.invoke(stocks_app, ['comparator', '--help'])
    assert result.exit_code == 0
    assert result.stdout != None

def test_command(requests_mock):
    symbols = ['GOOGL', 'AMZN']
    requests_mock.get(Endpoints.OVERVIEW, json=APIResponses.MS.OVERVIEW)
    requests_mock.get(Endpoints.INSTRUMENTS, json=APIResponses.MS.instruments(len(symbols)))
    requests_mock.get(Endpoints.AVG_VALUATION, json=APIResponses.MS.AVG_VALUATION)
    result = runner.invoke(stocks_app, ['comparator', *symbols])
    
    logger.debug(result.stdout)
    assert 'Loading' in result.stdout
    assert 'Finished' in result.stdout
    
    xlDoc = load_workbook(filename='comparation.xlsx', data_only=True)
    sheet: Worksheet = xlDoc.active
    initColumn = SpreadSheetComparation.initialColumn
    initRow = SpreadSheetComparation.initialRow
    rowMap = SpreadSheetComparation.rowMap
    
    iter = 0
    for cells in sheet.iter_cols(
        min_col=initColumn, 
        max_col=len(symbols) + initColumn - 1, 
        min_row=initRow, 
        max_row=40
    ):
        logger.info([ c.value for c in cells[:14] ])
        logger.info([ c.value for c in cells[23:35] ])
        
        assert cells[0].value == symbols[iter], 'the headers'
        
        assert cells[SpreadSheetComparation.rowMap['PER'] - initRow].value != None
        assert cells[SpreadSheetComparation.rowMap['PCF'] - initRow].value != None
        assert cells[SpreadSheetComparation.rowMap['PS'] - initRow].value != None
        assert cells[SpreadSheetComparation.rowMap['PBV'] - initRow].value != None

        assert cells[SpreadSheetComparation.rowMap['price'] - initRow].value != None

        iter += 1